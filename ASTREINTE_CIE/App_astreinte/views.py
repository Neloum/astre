# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Direction,Pole,Site,SousDirection
from .forms import DirectionForm,PoleForm,SiteForm,SousDirectionForm
from django.shortcuts import render, redirect
import re
import os
from django.conf import settings
from openpyxl import load_workbook
import pandas as pd
from django.shortcuts import render, redirect
from .models import Direction, SousDirection, Agent, UploadedFile
from .forms import FileUploadForm

from django.shortcuts import render, redirect, get_object_or_404
from .models import UploadedFile
import pandas as pd
import os
from django.http import HttpResponse



# Vue pour ajouter et afficher la liste des directions
def crud_view_directions(request):
    if request.method == 'POST':
        form = DirectionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('crud_view_directions')
    else:
        form = DirectionForm()

    directions = Direction.objects.all()  # Récupérer toutes les directions
    return render(request, 'directions.html', {'form': form, 'directions': directions})

# Vue pour modifier une direction
def modifier_direction(request, id):
    direction = get_object_or_404(Direction, id=id)
    if request.method == 'POST':
        form = DirectionForm(request.POST, instance=direction)
        if form.is_valid():
            form.save()
            return redirect('crud_view_directions')
    else:
        form = DirectionForm(instance=direction)
    return render(request, 'modifier_direction.html', {'form': form})

# Vue pour supprimer une direction
def supprimer_direction(request, id):
    direction = get_object_or_404(Direction, id=id)
    if request.method == 'POST':
        direction.delete()
        return redirect('crud_view_directions')
    return render(request, 'supprimer_direction.html', {'direction': direction})


def pole_crud_view(request, id=None):
    # Ajouter un pôle
    if request.method == 'POST' and 'add' in request.POST:
        form = PoleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('pole_crud_view')
    # Modifier un pôle
    elif request.method == 'POST' and 'edit' in request.POST:
        pole = get_object_or_404(Pole, id=id)
        form = PoleForm(request.POST, instance=pole)
        if form.is_valid():
            form.save()
            return redirect('pole_crud_view')
    # Supprimer un pôle
    elif request.method == 'POST' and 'delete' in request.POST:
        pole = get_object_or_404(Pole, id=id)
        pole.delete()
        return redirect('pole_crud_view')
    else:
        if id:
            pole = get_object_or_404(Pole, id=id)
            form = PoleForm(instance=pole)
        else:
            form = PoleForm()
    
    poles = Pole.objects.all()
    return render(request, 'poles_crud.html', {'form': form, 'poles': poles, 'editing': id is not None})


def site_crud_view(request, id=None):
    # Ajouter un site
    if request.method == 'POST' and 'add' in request.POST:
        form = SiteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('site_crud_view')
    # Modifier un site
    elif request.method == 'POST' and 'edit' in request.POST:
        site = get_object_or_404(Site, id=id)
        form = SiteForm(request.POST, instance=site)
        if form.is_valid():
            form.save()
            return redirect('site_crud_view')
    # Supprimer un site
    elif request.method == 'POST' and 'delete' in request.POST:
        site = get_object_or_404(Site, id=id)
        site.delete()
        return redirect('site_crud_view')
    else:
        if id:
            site = get_object_or_404(Site, id=id)
            form = SiteForm(instance=site)
        else:
            form = SiteForm()
    
    sites = Site.objects.all()
    return render(request, 'sites_crud.html', {'form': form, 'sites': sites, 'editing': id is not None})






def sousdirection_crud_view(request, id=None):
    sousdirection = None
    editing = False
    
    if id:
        sousdirection = get_object_or_404(SousDirection, id=id)
        editing = True

    if request.method == 'POST':
        # Suppression
        if 'delete' in request.POST and sousdirection:
            sousdirection.delete()
            return redirect('sousdirection_crud_view')

        # Modification ou ajout
        if editing:
            form = SousDirectionForm(request.POST, instance=sousdirection)
        else:
            form = SousDirectionForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('sousdirection_crud_view')

    else:
        if editing:
            form = SousDirectionForm(instance=sousdirection)
        else:
            form = SousDirectionForm()

    sousdirections = SousDirection.objects.all()
    return render(request, 'sousdirections_crud.html', {
        'form': form,
        'sousdirections': sousdirections,
        'editing': editing
    })





# Liste des e-mails et mots de passe autorisés
AUTHORIZED_CREDENTIALS = {
    'gracecodj01@gmail.com':'codjeau123',
    
}

def acceuil_view(request):
    return render(request, 'acceuil.html')



def login_view(request):
    if request.method == 'POST':
        email = request.POST['username']
        password = request.POST['password']

        if email in AUTHORIZED_CREDENTIALS and AUTHORIZED_CREDENTIALS[email] == password:
            return redirect('/acceuil/')  # Rediriger vers une URL absolue

        else:
            return render(request, 'login.html', {'error': 'Email ou mot de passe incorrect'})
    
    return render(request, 'login.html')




def upload_file(request):

    if request.method == 'POST':
        upload_form = FileUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            upload_form.save()
            # Rediriger après le téléchargement pour vider le formulaire
            return redirect('upload_file')
    else:
        upload_form = FileUploadForm()

    # Récupérer la liste des fichiers déjà uploadés
    files = UploadedFile.objects.all()
    # Si l'utilisateur veut supprimer un fichier
    if request.method == 'POST' and 'delete' in request.POST:
        file_id = request.POST.get('file_id')
        file_to_delete = get_object_or_404(UploadedFile, id=file_id)
        file_to_delete.delete()
        return redirect('upload_file')

    # Si l'utilisateur veut remplacer un fichier
    if request.method == 'POST' and 'replace' in request.POST:
        file_id = request.POST.get('file_id')
        new_file = request.FILES.get('replacement_file')
        file_to_replace = get_object_or_404(UploadedFile, id=file_id)
        file_to_replace.file = new_file
        file_to_replace.save()
        return redirect('upload_file')

    # Si l'utilisateur veut uploader un nouveau fichier
    if request.method == 'POST' and 'upload' in request.POST:
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('upload_file')

    # Si l'utilisateur veut extraire les données
    if request.method == 'POST' and 'extract' in request.POST:
        files = UploadedFile.objects.all()
        for file in files:
            df = pd.read_excel(file.file.path)  # Lire le fichier Excel
            # Logique pour traiter et extraire les données ici
        return redirect('upload_file')

    # Afficher les fichiers déjà uploadés
    uploaded_files = UploadedFile.objects.all()
    form = FileUploadForm()

    return render(request, 'upload.html', {
        'upload_form': form,
        'files': uploaded_files
    })




# Fonction pour nettoyer les numéros de téléphone
def clean_phone_number(phone):
    if phone:  # Si le téléphone n'est pas vide
        cleaned_phone = re.sub(r'\D', '', phone)  # Supprime tout ce qui n'est pas un chiffre
        return cleaned_phone if cleaned_phone else None  # Retourne None si le numéro est vide après nettoyage
    return None

# Fonction principale pour extraire les données depuis le fichier Excel
def extract_data_from_all_files(request):
    # Récupérer tous les fichiers uploadés
    uploaded_files = UploadedFile.objects.all()

    # Parcourir chaque fichier pour extraire les données
    for uploaded_file_instance in uploaded_files:
        # Charger le fichier Excel avec pandas
        df = pd.read_excel(uploaded_file_instance.file.path, skiprows=3)

        # Parcourir les lignes du fichier et enregistrer les données
        for index, row in df.iterrows():
            # Vérifier si la ligne n'est pas complètement vide
            if row.dropna().empty:
                print(f"Ligne {index} ignorée : vide")
                continue

            direction_nom = row.get('Direction')
            sous_direction_nom = row.get('Sous-Direction')
            nom_ag = row.get('Nom')
            prenom_ag = row.get('Prenom')
            telephone1 = str(row.get('Téléphone 1'))
            telephone2 = str(row.get('Téléphone 2'))

            # Affichage pour le débogage
            print(f"Nom: {nom_ag}, Prénom: {prenom_ag}, Téléphone 1: {telephone1}, Téléphone 2: {telephone2}")

            # Si le nom ou le prénom est manquant, ignorer la ligne
            if not nom_ag or not prenom_ag:
                print("Nom ou prénom manquant, ligne ignorée.")
                continue

            # Créer ou récupérer la Direction
            direction, _ = Direction.objects.get_or_create(nom_direc=direction_nom)

            # Si la sous-direction existe, la créer ou la récupérer, puis associer l'agent
            if sous_direction_nom and not pd.isna(sous_direction_nom):
                sous_direction, _ = SousDirection.objects.get_or_create(
                    lib_sous_direc=sous_direction_nom,
                    direction=direction
                )
                # Enregistrer l'agent sous la sous-direction, même sans numéro de téléphone
                Agent.objects.create(
                    nom_ag=nom_ag,
                    prenom_ag=prenom_ag,
                    tel_ag=telephone1 if telephone1 else None,
                    cell_ag=telephone2 if telephone2 else None,
                    sous_direction=sous_direction,
                    direction=direction
                )
            else:
                # Enregistrer l'agent directement sous la direction, même sans numéro de téléphone
                Agent.objects.create(
                    nom_ag=nom_ag,
                    prenom_ag=prenom_ag,
                    tel_ag=telephone1 if telephone1 else None,
                    cell_ag=telephone2 if telephone2 else None,
                    direction=direction
                )

    # Rediriger vers la page de téléchargement après extraction
    return redirect('upload_file')


from datetime import datetime, timedelta
import os
from django.http import HttpResponse
from openpyxl import load_workbook
from django.conf import settings
from .models import Direction, SousDirection, Agent

# Cellule où vous souhaitez ajouter la numérotation et la semaine dans le fichier Excel
NUM_ASTREINTE_CELL = 'K3'  # Cellule pour numéro d'astreinte
#SEMAINE_ASTREINTE_CELL = 'A2'  # Cellule pour la semaine de l'astreinte
DATE_DEBUT_CELL = 'O6'  # Cellule pour la date de début de la semaine
DATE_FIN_CELL = 'Y6'  # Cellule pour la date de fin de la semaine

# Fonction pour obtenir la semaine courante (du lundi au dimanche)
def get_semaine():
    today = datetime.today()
    start_of_week = today  # La date de départ est aujourd'hui
    end_of_week = start_of_week + timedelta(days=7)  # Dimanche
    return start_of_week, end_of_week

# Fonction pour générer le numéro de l'astreinte
def get_numero_astreinte(count):
    year = datetime.today().year
    numero = f"ASTREINTE CIE : N°{str(count).zfill(2)}.{year}"
    return numero

# Dictionnaire pour mapper les directions et sous-directions aux cellules spécifiques
cell_mapping = {
    "CDG": {"nom": "K13", "tel": "X13", "cell": "AG13"},
    "DIRECTION": {"nom": "K19", "tel": "X19", "cell": "AG19"},

    "DME": {"nom": "F25", "tel": "L25", "cell": "P25"},


    "DPE": {"nom": "F31", "tel": "L31", "cell": "P31"},
    "VRIDI1": {"nom": "F36", "tel": "L36", "cell": "P36"},
    "AYAME1": {"nom": "F38", "tel": "L38", "cell": "P38"},
    "KOSSOU": {"nom": "F40", "tel": "L40", "cell": "P40"},  
    "TAABO": {"nom": "F42", "tel": "L42", "cell": "P42"},
    "BUYO": {"nom": "F44", "tel": "L44", "cell": "P44"},
    "FAYE": {"nom": "F46", "tel": "L46", "cell": "P46"},


    "DCTET": {"nom": "F50", "tel": "L50", "cell": "P50"},
    "HELICO": {"nom": "F52", "tel": "L52", "cell": "P52"},
    "TLC/TRANS": {"nom": "F54", "tel": "L54", "cell": "P54"},
    "RADIO": {"nom": "F56", "tel": "L56", "cell": "P56"},
    "SD_MAINT": {"nom": "F58", "tel": "L58", "cell": "P58"},
    "DRTET_ABJ": {"nom": "F60", "tel": "L60", "cell": "P60"},
    "DRTET_ABE": {"nom": "F62", "tel": "L62", "cell": "P62"},
    "DRTET_BKE": {"nom": "F64", "tel": "L64", "cell": "P64"},
    "DRTET_KRO": {"nom": "F66", "tel": "L66", "cell": "P66"},
    "DRTET_MAN": {"nom": "F68", "tel": "L68", "cell": "P68"},
    "DRTET_SOU": {"nom": "F70", "tel": "L70", "cell": "P70"},

    "DST": {"nom": "F74", "tel": "L74", "cell": "P74"},
    "CME": {"nom": "F78", "tel": "L78", "cell": "P78"},

    "DGA DPSC": {"nom": "AA25", "tel": "AG25", "cell": "AK25"},

    "DCRD": {"nom": "AA31", "tel": "AG31", "cell": "AK31"},

    "MCE/AB": {"nom": "AA36", "tel": "AG36", "cell": "AK36"},
    "HTA/AB": {"nom": "AA38", "tel": "AG38", "cell": "AK38"},
    "DAMI": {"nom": "AA40", "tel": "AG40", "cell": "AK40"},
    "DRAN": {"nom": "AA42", "tel": "AG42", "cell": "AK42"},
    "DRAS": {"nom": "AA44", "tel": "AG44", "cell": "AK44"},
    "DRYOP": {"nom": "AA46", "tel": "AG46", "cell": "AK46"},
    "DRABO": {"nom": "AA48", "tel": "AG48", "cell": "AK48"},
    "DRE": {"nom": "AA50", "tel": "AG50", "cell": "AK50"},
    "DRBC": {"nom": "AA52", "tel": "AG52", "cell": "AK52"},
    "DRSE": {"nom": "AA54", "tel": "AG54", "cell": "AK54"},
    "DRC": {"nom": "AA56", "tel": "AG56", "cell": "AK56"},
    "DRCO": {"nom": "AA58", "tel": "AG58", "cell": "AK58"},
    "DRSO": {"nom": "AA60", "tel": "AG60", "cell": "AK60"},
    "DRN": {"nom": "AA62", "tel": "AG62", "cell": "AK62"},
    "DRO": {"nom": "AA64", "tel": "AG64", "cell": "AK64"},
    "DRCS": {"nom": "AA66", "tel": "AG66", "cell": "AK66"},
    "DRLO": {"nom": "AA68", "tel": "AG68", "cell": "AK68"},
    "DEP": {"nom": "AA70", "tel": "AG70", "cell": "AK70"},


    "DMT": {"nom": "AA74", "tel": "AG74", "cell": "AK74"},
    "DPS": {"nom": "AA78", "tel": "AG78", "cell": "AK78"},

   

    # Ajoute ici les autres directions/sous-directions et leurs cellules correspondantes
}


def generate_astreinte_and_download(request):
    # Charger le fichier modèle
    template_path = os.path.join(settings.MEDIA_ROOT, 'files', 'ASTREINTE_CIE.xlsx')
    wb = load_workbook(template_path)
    sheet = wb.active


    # Générer la numérotation d'astreinte (utiliser un compteur dynamique)
    count = 1  # Exemple de compteur
    numero_astreinte = get_numero_astreinte(count)

    # Obtenir la semaine courante
    start_of_week, end_of_week = get_semaine()
    #semaine_astreinte = f"Semaine du {start_of_week.strftime('%d/%m/%Y')} au {end_of_week.strftime('%d/%m/%Y')}"
    debut_semaine_str = start_of_week.strftime('%d_%m_%Y')
    fin_semaine_str = end_of_week.strftime('%d_%m_%Y')

    # Insérer la numérotation et la semaine dans les cellules spécifiées
    sheet[NUM_ASTREINTE_CELL] = numero_astreinte
    #sheet[SEMAINE_ASTREINTE_CELL] = semaine_astreinte
    sheet[DATE_DEBUT_CELL] = start_of_week.strftime('%d/%m/%Y')
    sheet[DATE_FIN_CELL] = end_of_week.strftime('%d/%m/%Y')

    # Parcourir chaque direction/sous-direction
    for entity_name, cells in cell_mapping.items():
        try:
            # Vérifier si c'est une direction
            direction = Direction.objects.filter(nom_direc=entity_name).first()

            # Sinon, chercher la sous-direction
            if direction:
                agent = Agent.objects.filter(direction=direction).first()
            else:
                sous_direction = SousDirection.objects.filter(lib_sous_direc=entity_name).first()
                if sous_direction:
                    agent = Agent.objects.filter(sous_direction=sous_direction).first()
                    direction = sous_direction.direction
                else:
                    continue

            if agent:
                nom_complet = f"{agent.nom_ag} {agent.prenom_ag}"
                tel_ag = agent.tel_ag
                cell_ag = agent.cell_ag

                # Remplir les cellules correspondantes dans le fichier Excel
                sheet[cells['nom']] = nom_complet
                sheet[cells['tel']] = tel_ag
                sheet[cells['cell']] = cell_ag
        except Exception as e:
            print(f"Erreur lors de la recherche de {entity_name} : {e}")

     # Sauvegarder le fichier modifié dans un nouveau chemin
    chemin_fichier_modifie = os.path.join(settings.MEDIA_ROOT, 'files', f'astreinte_final_semaine_du_{debut_semaine_str}_au_{fin_semaine_str}.xlsx')
    wb.save(chemin_fichier_modifie)

    # Générer le nom de fichier avec la plage de dates
    nom_fichier = f"astreinte_final_semaine_du_{debut_semaine_str}_au_{fin_semaine_str}.xlsx"

    # Télécharger le fichier Excel modifié avec le nom dynamique
    with open(chemin_fichier_modifie, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response